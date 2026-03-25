"""
Export Manager — Phase 2
========================
Exports the CPM schedule to:
  • Excel (.xlsx)  — formatted table with CPM columns + critical-path highlighting
  • PDF   (.pdf)   — reportlab-based schedule report with Gantt-style summary

Dependencies (added to requirements.txt):
    openpyxl>=3.1
    reportlab>=4.0
"""
from __future__ import annotations

import io
import os
from datetime import date, timedelta
from typing import Dict, List, Optional

from activity import Activity


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(
    activities: Dict[str, Activity],
    filepath: str,
    project_name: str = "CPM Schedule",
    start_date: Optional[date] = None,
) -> None:
    """
    Write the schedule to an Excel workbook.

    Parameters
    ----------
    activities  : {id: Activity}
    filepath    : destination .xlsx path
    project_name: written as a title in cell A1
    start_date  : if supplied, date columns (Start Date, Finish Date) are computed
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for Excel export.  Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CPM Schedule"

    # ---- colour palette ----
    RED_FILL   = PatternFill("solid", fgColor="C04040")
    BLUE_FILL  = PatternFill("solid", fgColor="3A7CA5")
    GREY_FILL  = PatternFill("solid", fgColor="D0D8E4")
    HDR_FILL   = PatternFill("solid", fgColor="1E2530")
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

    thin = Side(border_style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    white_bold  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    dark_bold   = Font(name="Calibri", bold=True, color="1E2530", size=14)
    dark_normal = Font(name="Calibri", color="202020", size=10)
    red_bold    = Font(name="Calibri", bold=True, color="C04040", size=10)

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    # ---- Title row ----
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = project_name
    title_cell.font  = dark_bold
    title_cell.alignment = left
    title_cell.fill = PatternFill("solid", fgColor="E8E8E8")
    ws.row_dimensions[1].height = 28

    if start_date:
        ws["N1"] = f"Project Start: {start_date.strftime('%d %b %Y')}"
        ws["N1"].font = Font(name="Calibri", italic=True, color="505050", size=10)

    # ---- Headers ----
    has_dates = start_date is not None
    headers = [
        "ID", "Activity Name", "Duration", "Predecessors", "Resource",
        "ES", "EF", "LS", "LF", "Total Float", "Free Float", "Critical",
    ]
    if has_dates:
        headers += ["Early Start Date", "Early Finish Date"]

    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
        cell.fill      = HDR_FILL
        cell.font      = white_bold
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[2].height = 20

    # ---- Data rows ----
    for row_idx, act in enumerate(activities.values(), start=3):
        is_crit = act.is_critical
        row_fill = RED_FILL if is_crit else (GREY_FILL if row_idx % 2 == 0 else WHITE_FILL)
        font = red_bold if is_crit else dark_normal

        row_data = [
            act.id,
            act.name,
            act.duration,
            ",".join(act.predecessors) or "—",
            act.resource or "",
            act.ES,
            act.EF,
            act.LS,
            act.LF,
            act.total_float,
            act.free_float,
            "★ CRITICAL" if is_crit else "",
        ]
        if has_dates:
            es_date = start_date + timedelta(days=act.ES)
            ef_date = start_date + timedelta(days=act.EF)
            row_data += [es_date.strftime("%d %b %Y"), ef_date.strftime("%d %b %Y")]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.font      = font
            cell.border    = border
            cell.alignment = left if col_idx == 2 else center

        ws.row_dimensions[row_idx].height = 18

    # ---- Column widths ----
    col_widths = [8, 30, 10, 18, 16, 7, 7, 7, 7, 12, 12, 12, 18, 18]
    for i, w in enumerate(col_widths[:len(headers)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Summary sheet ----
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Project Summary"
    ws2["A1"].font = dark_bold
    total_dur = max((a.EF for a in activities.values()), default=0)
    critical  = [a.id for a in activities.values() if a.is_critical]
    ws2["A3"] = "Total Activities"
    ws2["B3"] = len(activities)
    ws2["A4"] = "Project Duration (days)"
    ws2["B4"] = total_dur
    ws2["A5"] = "Critical Activities"
    ws2["B5"] = len(critical)
    ws2["A6"] = "Critical Path"
    ws2["B6"] = " → ".join(critical) or "N/A"
    if has_dates:
        ws2["A7"] = "Project Start"
        ws2["B7"] = start_date.strftime("%d %b %Y")
        ws2["A8"] = "Project End (Early)"
        ws2["B8"] = (start_date + timedelta(days=total_dur)).strftime("%d %b %Y")

    for row in ws2.iter_rows(min_row=3, max_row=8, min_col=1, max_col=2):
        for cell in row:
            cell.font   = dark_normal
            cell.border = border

    wb.save(filepath)


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------

def export_to_pdf(
    activities: Dict[str, Activity],
    filepath: str,
    project_name: str = "CPM Schedule",
    start_date: Optional[date] = None,
) -> None:
    """
    Write a schedule report to PDF using reportlab.

    Parameters
    ----------
    activities  : {id: Activity}
    filepath    : destination .pdf path
    project_name: title shown on the report
    start_date  : optional; enables date columns in the table
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable
        )
    except ImportError:
        raise ImportError("reportlab is required for PDF export.  Run: pip install reportlab")

    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        filepath,
        pagesize=page_size,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#1E2530"),
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "Sub",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#606060"),
        spaceAfter=10,
    )

    story = []

    # ---- Title ----
    story.append(Paragraph(project_name, title_style))
    total_dur = max((a.EF for a in activities.values()), default=0)
    critical  = [a.id for a in activities.values() if a.is_critical]
    sub_text = (
        f"Generated: {date.today().strftime('%d %b %Y')}   |   "
        f"Activities: {len(activities)}   |   "
        f"Duration: {total_dur} days   |   "
        f"Critical Path: {' → '.join(critical) or 'N/A'}"
    )
    if start_date:
        finish = start_date + timedelta(days=total_dur)
        sub_text += (
            f"   |   Start: {start_date.strftime('%d %b %Y')}"
            f"   |   Finish: {finish.strftime('%d %b %Y')}"
        )
    story.append(Paragraph(sub_text, sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#C0C0C0")))
    story.append(Spacer(1, 8))

    # ---- Table ----
    has_dates = start_date is not None
    headers = ["ID", "Activity Name", "Dur", "Predecessors", "Resource",
               "ES", "EF", "LS", "LF", "TF", "FF", "Critical"]
    if has_dates:
        headers += ["Early Start", "Early Finish"]

    table_data = [headers]
    for act in activities.values():
        row = [
            act.id,
            act.name,
            str(act.duration),
            ",".join(act.predecessors) or "—",
            act.resource or "",
            str(act.ES), str(act.EF), str(act.LS), str(act.LF),
            str(act.total_float), str(act.free_float),
            "★ YES" if act.is_critical else "",
        ]
        if has_dates:
            row += [
                (start_date + timedelta(days=act.ES)).strftime("%d %b %Y"),
                (start_date + timedelta(days=act.EF)).strftime("%d %b %Y"),
            ]
        table_data.append(row)

    col_widths_cm = [1.4, 5.5, 1.2, 3.0, 2.5,
                     1.2, 1.2, 1.2, 1.2, 1.2, 1.2, 1.8]
    if has_dates:
        col_widths_cm += [2.8, 2.8]
    col_widths_pt = [w * cm for w in col_widths_cm]

    tbl = Table(table_data, colWidths=col_widths_pt, repeatRows=1)

    # Build per-row styling
    style_cmds = [
        # Header row
        ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#1E2530")),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0), 9),
        ("ALIGN",        (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING",(0, 0), (-1, 0), 6),
        ("TOPPADDING",   (0, 0), (-1, 0), 6),
        # Body
        ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",     (0, 1), (-1, -1), 8),
        ("ALIGN",        (0, 1), (-1, -1), "CENTER"),
        ("ALIGN",        (1, 1), (1,  -1), "LEFT"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1),
         [colors.HexColor("#FFFFFF"), colors.HexColor("#F0F4F8")]),
        ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#C0C0C0")),
        ("TOPPADDING",   (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 1), (-1, -1), 4),
    ]

    # Highlight critical rows
    for i, act in enumerate(activities.values(), start=1):
        if act.is_critical:
            style_cmds.append(
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFE8E8"))
            )
            style_cmds.append(
                ("TEXTCOLOR", (0, i), (-1, i), colors.HexColor("#B03040"))
            )
            style_cmds.append(
                ("FONTNAME", (0, i), (-1, i), "Helvetica-Bold")
            )

    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)

    doc.build(story)
